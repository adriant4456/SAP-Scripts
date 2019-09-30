#Check in gui automation
import pyautogui
import win32gui
import win32con
import time
import datetime
import shutil
import os
import openpyxl as pyxl

def mbom(material):
    pyautogui.PAUSE = 0.5
    handle = WindowEnumerate('SAP Easy Access')
    print(handle)
    win32gui.ShowWindow(handle, win32con.SW_MAXIMIZE)
    win32gui.SetForegroundWindow(handle)
    time.sleep(2)
    pyautogui.hotkey('esc')
    pyautogui.typewrite('CS12')
    pyautogui.hotkey('enter')
    wait_window_focus('Explode BOM: Multi-Level BOM: Initial Screen')
    pyautogui.typewrite(str(material))
    pyautogui.hotkey('tab')
    pyautogui.typewrite('3010')
    pyautogui.hotkey('f8')
    wait_window_focus('Display Multilevel BOM')
    material = str(material)
    if not os.path.isdir(f"C:\\Users\\atai\\Desktop\\{material}"):
        os.mkdir(f"C:\\Users\\atai\\Desktop\\{material}")
    time.sleep(4)
    pyautogui.hotkey('ctrl', 'shift', 'f9')
    time.sleep(1)
    pyautogui.hotkey('down')
    pyautogui.hotkey('enter')
    time.sleep(1)
    pyautogui.hotkey('delete')
    pyautogui.typewrite(str(material))
    pyautogui.typewrite('.XLS')    
    pyautogui.hotkey('up')
    pyautogui.hotkey('ctrl', 'a')
    pyautogui.hotkey('delete')
    pyautogui.typewrite(f"C:\\Users\\atai\\Desktop\\{material}")
    pyautogui.hotkey('enter')
    wait_window_focus('Display Multilevel BOM')
    pyautogui.hotkey('f12')
    pyautogui.hotkey('f12')

def plot_struc(material):
    pyautogui.PAUSE = 0.5
    handle = WindowEnumerate('SAP Easy Access')
    print(handle)
    win32gui.ShowWindow(handle, win32con.SW_MAXIMIZE)
    win32gui.SetForegroundWindow(handle)
    time.sleep(2)
    pyautogui.hotkey('esc')
    pyautogui.typewrite('z1000039')
    pyautogui.hotkey('enter')
    wait_window_focus('BOM Plot')
    pyautogui.hotkey('tab')
    pyautogui.hotkey('tab')
    pyautogui.typewrite(str(material))
    pyautogui.hotkey('tab')
    pyautogui.typewrite('3010')
    pyautogui.hotkey('tab')
    pyautogui.hotkey('down')
    pyautogui.hotkey('down')
    pyautogui.hotkey('tab')
    pyautogui.hotkey('tab')
    pyautogui.hotkey('tab')
    pyautogui.typewrite('1')
    pyautogui.hotkey('tab')
    pyautogui.hotkey('space')
    pyautogui.hotkey('f8')
    while True:
        modified_list = []
        out = False
        for i in os.listdir('C:\\Temp'):
            mtime = os.path.getmtime('C:\\Temp\\' + i)
            current_time = time.time()
            if (current_time - mtime) < 60: #check if file was created in last 1 mins
                modified_list.append(i)
            print(modified_list)
        for i in modified_list:
            if 'isbom' in i:
                print('should break')
                out = True
                break
        if out:
            break
        time.sleep(0.5)
    time.sleep(1)
    if not pyautogui.confirm('Continue to DocStructure?') == 'OK':
        exit()
    handle = WindowEnumerate('BOM Plot')
    print(handle)
    win32gui.ShowWindow(handle, win32con.SW_MAXIMIZE)
    win32gui.SetForegroundWindow(handle)
    pyautogui.hotkey('f12')
    pyautogui.hotkey('f12')
    modified_list = []
    for i in os.listdir('C:\\Temp'):
        mtime = os.path.getmtime('C:\\Temp\\' + i)
        current_time = time.time()
        if (current_time - mtime) < 120: #check if file was created in last 2 mins
            if "tmp" not in i:
                modified_list.append(i)
    #find name of file
    try:
        folder_name = modified_list[1].split('.')[0]
    except:
        print('Can\'t find files')
        exit()
    folder_dir = 'C:\\Users\\atai\\Desktop\\' + folder_name
    if not os.path.isdir(f"C:\\Users\\atai\\Desktop\\{material}"):
        os.mkdir(folder_dir)
    for i in modified_list:
        shutil.move('C:\\Temp\\' + i, folder_dir)
    pyautogui.hotkey('esc')
    pyautogui.typewrite('z1000082')
    pyautogui.hotkey('enter')
    wait_window_focus('Inventor Document Structure Check')
    pyautogui.typewrite(str(material))
    pyautogui.hotkey('f8')
    while True:
        modified_list = []
        out = False
        for i in os.listdir('C:\\Temp'):
            mtime = os.path.getmtime('C:\\Temp\\' + i)
            current_time = time.time()
            if (current_time - mtime) < 60: #check if file was created in last 1 mins
                modified_list.append(i)
            print(modified_list)
        for i in modified_list:
            if 'ErrorLog' in i:
                print('should break')
                out = True
                break
        if out:
            break
        time.sleep(0.5)
    time.sleep(1)
    modified_list = []
    for i in os.listdir('C:\\Temp'):
        mtime = os.path.getmtime('C:\\Temp\\' + i)
        current_time = time.time()
        if (current_time - mtime) < 120: #check if file was created in last 2 mins
            modified_list.append(i)
    for i in modified_list:
        if 'ErrorLog' in i:
            shutil.move('C:\\Temp\\' + i, folder_dir)
    time.sleep(1)
    if not pyautogui.confirm('Continue to SAP?') == 'OK':
        exit()
    handle = WindowEnumerate('INVENTOR DOCUMENT STRUCTURE')
    print(handle)
    win32gui.ShowWindow(handle, win32con.SW_MAXIMIZE)
    win32gui.SetForegroundWindow(handle)
    pyautogui.moveTo(360,206)
    pyautogui.click()
    time.sleep(1)
    pyautogui.hotkey('down')
    pyautogui.hotkey('enter')
    pyautogui.hotkey('enter')
    wait_window_focus('Save As')
    time.sleep(1)
    pyautogui.typewrite(str(material) + '.MHTML')
    while True:
        out = False
        for i in os.listdir('C:\\Users\\atai\\Documents\\SAP\\SAP GUI'):
            if str(material) in i:
                out = True
                break
        if out:
            break
    shutil.move('C:\\Users\\atai\\Documents\\SAP\\SAP GUI\\' + str(material) + '.MHTML', folder_dir)

    
#timesheet entry

def sap_timesheet():
    #launch timesheet in excel
    os.system("start EXCEL>EXE C:\\Users\\active_timesheet.xlsx")
    waitforwindow('active_timesheet.xlsx')
    #bring to front
    handle = WindowEnumerate('active_timesheet.xlsx')
    print(handle)
    win32gui.ShowWindow(handle, win32con.SW_MAXIMIZE)
    win32gui.SetForegroundWindow(handle)
    #confirm timesheet is correct
    if not pyautogui.confirm('Timesheet is correct? Continue to SAP?') == 'OK':
        exit()

    #goto SAP create network screen
    pyautogui.PAUSE = 0.5
    handle = WindowEnumerate('SAP Easy Access')
    print(handle)
    win32gui.ShowWindow(handle, win32con.SW_MAXIMIZE)
    win32gui.SetForegroundWindow(handle)
    time.sleep(2)
    wait_window_focus('SAP Easy Access')
    pyautogui.hotkey('esc')
    pyautogui.typewrite('CN25')
    pyautogui.hotkey('enter')
    wait_window_focus('Create Network Confirmation: Initial Screen')
    #access timesheet.xlsx
    try:
        wb = pyxl.load_workbook('C:\\Users\\active_timesheet.xlsx')
        sheet= wb.active
    except NameError:
        raise

    print('accessed timesheet')
    
    #iterate through time entries
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=sheet.max_column,max_row=sheet.max_row):
        for cell in row:
            if cell.value is None:
                pass
            else:
                time_entry = cell.value
                posting_date = sheet.cell(row=1, column = cell.column).value
                #format posting date
                posting_date = f"{posting_date.day}.{posting_date.month}.{posting_date.year}"
                #get and format network and activity
                wbs = sheet.cell(row=cell.row, column = 2).value
                [network,activity] = wbs.split('/')
                
                print('set network and activity')
                
                #log to SAP
                log_network_to_SAP(network, activity, posting_date, time_entry)

             
        
        



    
def log_network_to_SAP(network,activity,posting_date,time_entry):
    #check for confirm network screen
    wait_window_focus('Create Network Confirmation: Initial Screen')
    #prefill network and activity number and confirm
    pyautogui.typewrite(network)
    pyautogui.hotkey('tab')
    pyautogui.typewrite(activity)
    pyautogui.hotkey('enter')
    time.sleep(1)
    pyautogui.hotkey('enter')
    wait_window_focus('Create Network Confirmation: Actual Data')
    #tab to 'posting date' and enter
    pyautogui.PAUSE = 0.1
    for i in range(11):
        pyautogui.hotkey('tab')
    pyautogui.PAUSE = 0.5
    pyautogui.typewrite(posting_date)

    #tab to 'hours' and enter
    pyautogui.PAUSE = 0.1
    for i in range(12):
        pyautogui.hotkey('tab')
    pyautogui.PAUSE = 0.5
    pyautogui.typewrite(str(time_entry))

    #wait for user to confirm timesheet entry, check for create network confirmation window
    wait_window_focus('Create Network Confirmation: Initial Screen')
    time.sleep(1)
    #goes back to create network screen
    
def waitforwindow(window):
    while True:
        if WindowEnumerate(window):
            break
        time.sleep(1)

def wait_window_focus(window):
    #check for "window" in the title of focussed window every second
    while True:
        if window in win32gui.GetWindowText(win32gui.GetForegroundWindow()):
            print('done waiting')
            break
        time.sleep(1)
    

def WindowEnumerate(window):
    top_windows = []
    win32gui.EnumWindows(windowEnumerationHandler, top_windows)
    for i in top_windows:
        if window in i[1]:
            print(i)
            return i[0]
    return False

def windowEnumerationHandler(hwnd, top_windows):
    top_windows.append((hwnd, win32gui.GetWindowText(hwnd)))

